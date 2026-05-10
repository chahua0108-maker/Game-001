#!/usr/bin/env python3
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path("outputs/research/vampire-crawlers")
OUT = BASE / "vampire_crawlers_research.xlsx"


THEMES = {
    "addictive_s爽": [
        "addictive",
        "addicting",
        "addicted",
        "hooked",
        "one more",
        "dopamine",
        "satisfying",
        "fun",
        "爽",
        "上头",
        "停不下来",
        "停不下來",
        "神",
        "好玩",
    ],
    "boring_repetitive": [
        "boring",
        "bored",
        "repetitive",
        "repeat",
        "chore",
        "grind",
        "grindy",
        "枯燥",
        "重复",
        "重複",
        "无聊",
        "無聊",
        "刷",
        "肝",
    ],
    "deck_combo_build": [
        "deck",
        "card",
        "combo",
        "build",
        "synergy",
        "mana",
        "wildcard",
        "gem",
        "卡",
        "牌",
        "连击",
        "連擊",
        "构筑",
        "構築",
        "宝石",
        "寶石",
        "万能宝石",
    ],
    "ui_readability": [
        "ui",
        "interface",
        "readability",
        "readable",
        "menu",
        "controller",
        "handheld",
        "switch",
        "steam deck",
        "界面",
        "菜单",
        "菜單",
        "手柄",
        "掌机",
        "掌機",
    ],
    "slow_progression": [
        "slow",
        "too long",
        "takes",
        "hours",
        "unlock",
        "progression",
        "gold",
        "money",
        "慢",
        "解锁",
        "解鎖",
        "金币",
        "金幣",
    ],
    "bug_performance": [
        "bug",
        "crash",
        "fps",
        "performance",
        "lag",
        "stutter",
        "卡顿",
        "卡頓",
        "崩溃",
        "崩潰",
        "bug",
    ],
    "vampire_survivors_identity": [
        "vampire survivors",
        "survivors",
        "poncle",
        "vs",
        "吸血鬼幸存者",
        "吸血鬼倖存者",
        "ヴァンサバ",
    ],
    "price_value": [
        "price",
        "worth",
        "value",
        "cheap",
        "game pass",
        "$",
        "价格",
        "價格",
        "值",
        "便宜",
    ],
}


MEDIA_ROWS = [
    ["PC Gamer", "2026-04-20", "Review", "50/100 / negative outlier", "即时升级刺激仍成立", "前期路径单一、重复、永久升级像门槛", "第一局必须给多路线决策"],
    ["Windows Central", "2026-04-20", "Review", "Strong positive", "短局循环、解锁、平台适配", "前期更慢，掌机可读性问题", "手柄/掌机 UI 早测"],
    ["Shacknews", "2026-02-25", "Preview", "Positive", "快节奏、简化、新鲜感", "地牢可能只是工具化容器", "地图需要风险路径"],
    ["VGC", "2026-04-25", "Review", "4/5", "主动操作感和 build 变化", "内容成熟度仍有限", "证明主动操作而非换皮"],
    ["GameSpot", "2026-04-20", "Review", "8/10", "熟悉与新鲜平衡，卡牌连锁上头", "完整爽点 15-20 小时后才成形", "10 分钟 demo 必须爆发"],
    ["Destructoid", "2026-04-20", "Review", "9/10", "简单、上头、满足、解锁量足", "早期经济偏慢，战斗深度有限", "经济曲线服务实验"],
    ["Nintendo Life", "2026-04-20", "Review", "8/10", "组合、进化、解锁、掌机模式", "bug，升级信息不足", "奖励解释未来收益"],
    ["Push Square", "2026-04-20", "Review", "8/10", "核心循环、快节奏、进度感", "移动僵硬，手柄菜单繁琐", "移动和菜单不是次要体验"],
    ["Game Informer", "2026-05", "Review", "Positive", "灯光声音按钮构成感官系统", "刺激可能接近注意力轰炸", "爽感是反馈编排"],
    ["Pocket Tactics", "2026-04-20", "Review", "9/10", "hypnotic，Village hub 融入成就", "小性能问题", "Hub 是长线目标板"],
    ["Game8", "2026-04", "Review", "84", "混合类型独特", "需核对更细缺点", "非欧美口径补充"],
    ["Forbes", "2026-04-27", "Review", "Positive", "保留 frantic gameplay 且有新转法", "需核对全文", "原创版本要有自己的 frantic 来源"],
    ["The Outerhaven", "2026-04-20", "Review", "90", "build 实验和 one more run", "不一定适合所有玩家", "定义目标玩家"],
    ["4Gamer", "2026-04-23", "Feature", "Positive", "把爽快感压缩进 deckbuilding", "长期节奏未完全验证", "组合爆炸需 UI 支撑"],
    ["Famitsu", "2026-03-20", "News", "Info", "破坏性构筑和抽牌过载", "信息负荷可能高", "卡牌过载要有整理工具"],
    ["Dengeki", "2026-04-20", "Feature", "Positive", "混乱回合制卡牌和音乐卖点", "推荐文不拆缺点", "音乐可放大机制记忆点"],
]


def ts_to_iso(value):
    if not value:
        return ""
    try:
        return datetime.fromtimestamp(int(value), timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def load_reviews():
    path = BASE / "steam_reviews_merged_flat.csv"
    if not path.exists():
        path = BASE / "steam_reviews_flat.csv"
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["voted_up_bool"] = row.get("voted_up") == "True"
        row["created_iso"] = ts_to_iso(row.get("timestamp_created"))
        row["playtime_hours_at_review"] = round((int(row.get("author.playtime_at_review") or 0) / 60), 2)
        row["review_len"] = len(row.get("review") or "")
        text = (row.get("review") or "").lower()
        matched = []
        for theme, terms in THEMES.items():
            if any(term.lower() in text for term in terms):
                matched.append(theme)
        row["themes"] = ";".join(matched)
    return rows


def load_youtube():
    summary_path = BASE / "youtube/youtube_collection_summary.json"
    if not summary_path.exists():
        return [], []
    data = json.loads(summary_path.read_text(encoding="utf-8"))
    return data.get("videos") or [], data.get("segments") or []


def load_steam_endpoint_summary():
    for path in [
        BASE / "rescue/steam_review_remaining_language_raw.json",
        BASE / "steam_reviews_by_language_raw.json",
    ]:
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        summary = data.get("language_all_query_summary")
        if summary:
            return summary
    return {}


def set_table_style(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def autofit(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def append_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    set_table_style(ws)
    autofit(ws)
    return ws


def main():
    reviews = load_reviews()
    videos, segments = load_youtube()
    steam_summary = load_steam_endpoint_summary()
    wb = Workbook()
    wb.remove(wb.active)

    positive = sum(1 for r in reviews if r["voted_up_bool"])
    negative = len(reviews) - positive
    languages = Counter(r["language"] or "unknown" for r in reviews)
    themes = Counter(theme for r in reviews for theme in r["themes"].split(";") if theme)
    theme_pos = defaultdict(lambda: [0, 0])
    for r in reviews:
        for theme in r["themes"].split(";"):
            if theme:
                theme_pos[theme][0 if r["voted_up_bool"] else 1] += 1

    summary_rows = [
        ["数据状态", "Steam 用户评测为分语言合并数据集", f"已抓 {len(reviews)} 条；主要语言与已发现极小语种已合并"],
        ["Steam endpoint total_reviews", steam_summary.get("total_reviews", ""), "公开 endpoint 会随新评论实时变化"],
        ["Steam endpoint total_positive/negative", f"{steam_summary.get('total_positive', '')} / {steam_summary.get('total_negative', '')}", ""],
        ["Steam 数据源", "steam_reviews_merged_flat.csv" if (BASE / "steam_reviews_merged_flat.csv").exists() else "steam_reviews_flat.csv", ""],
        ["时间范围", min(r["created_iso"] for r in reviews if r["created_iso"]), max(r["created_iso"] for r in reviews if r["created_iso"])],
        ["推荐/不推荐", positive, negative],
        ["推荐率", round(positive / len(reviews), 4) if reviews else "", ""],
        ["YouTube 视频数", len(videos), ""],
        ["YouTube 短片段数", len(segments), ""],
        ["媒体来源数", len(MEDIA_ROWS), ""],
        ["核心限制", "Steam 全局 cursor 不稳定，已改用分语言桶；YouTube 部分字幕/元数据受 429 限制", ""],
    ]
    ws = append_sheet(wb, "Summary", ["Metric", "Value", "Notes"], summary_rows)

    chart = PieChart()
    chart.title = "Steam Sentiment in Partial Dataset"
    pie_data_start = ws.max_row + 2
    ws.cell(pie_data_start, 1, "Sentiment")
    ws.cell(pie_data_start, 2, "Count")
    ws.cell(pie_data_start + 1, 1, "Positive")
    ws.cell(pie_data_start + 1, 2, positive)
    ws.cell(pie_data_start + 2, 1, "Negative")
    ws.cell(pie_data_start + 2, 2, negative)
    data = Reference(ws, min_col=2, min_row=pie_data_start, max_row=pie_data_start + 2)
    labels = Reference(ws, min_col=1, min_row=pie_data_start + 1, max_row=pie_data_start + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, "E2")

    append_sheet(
        wb,
        "Steam Raw Reviews",
        [
            "recommendationid",
            "language",
            "voted_up",
            "created_iso",
            "playtime_hours_at_review",
            "steam_purchase",
            "received_for_free",
            "votes_up",
            "votes_funny",
            "weighted_vote_score",
            "comment_count",
            "review_len",
            "themes",
            "review",
        ],
        [
            [
                r.get("recommendationid"),
                r.get("language"),
                r.get("voted_up"),
                r.get("created_iso"),
                r.get("playtime_hours_at_review"),
                r.get("steam_purchase"),
                r.get("received_for_free"),
                r.get("votes_up"),
                r.get("votes_funny"),
                r.get("weighted_vote_score"),
                r.get("comment_count"),
                r.get("review_len"),
                r.get("themes"),
                r.get("review"),
            ]
            for r in reviews
        ],
    )

    lang_rows = [[lang, count, round(count / len(reviews), 4)] for lang, count in languages.most_common()]
    append_sheet(wb, "Language Analysis", ["Language", "Reviews", "Share"], lang_rows)

    theme_rows = []
    for theme, count in themes.most_common():
        pos, neg = theme_pos[theme]
        theme_rows.append([theme, count, pos, neg, round(pos / count, 4) if count else ""])
    ws_theme = append_sheet(wb, "Theme Taxonomy", ["Theme", "Mentions", "Positive", "Negative", "Positive Rate"], theme_rows)
    if theme_rows:
        chart = BarChart()
        chart.title = "Theme Mentions"
        chart.y_axis.title = "Mentions"
        chart.x_axis.title = "Theme"
        data = Reference(ws_theme, min_col=2, min_row=1, max_row=min(len(theme_rows) + 1, 12))
        cats = Reference(ws_theme, min_col=1, min_row=2, max_row=min(len(theme_rows) + 1, 12))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_theme.add_chart(chart, "G2")

    media_headers = ["Media", "Date", "Type", "Score/Conclusion", "Positive", "Negative", "Design Question"]
    append_sheet(wb, "Media Reviews", media_headers, MEDIA_ROWS)

    video_rows = []
    for v in videos:
        video_rows.append(
            [
                v.get("label"),
                v.get("id"),
                v.get("title"),
                v.get("channel"),
                v.get("url"),
                v.get("duration_string") or v.get("duration"),
                v.get("upload_date"),
                v.get("view_count"),
                v.get("like_count"),
                v.get("comment_count_reported"),
                v.get("comments_collected"),
                v.get("info_json"),
                "; ".join(v.get("subtitle_files") or []),
                v.get("metadata_returncode"),
            ]
        )
    append_sheet(
        wb,
        "Video Materials",
        ["Label", "Video ID", "Title", "Channel", "URL", "Duration", "Upload Date", "Views", "Likes", "Reported Comments", "Collected Comments", "Info JSON", "Subtitle Files", "Metadata Return Code"],
        video_rows,
    )

    segment_rows = []
    for s in segments:
        segment_rows.append(
            [
                s.get("label"),
                s.get("video_id"),
                s.get("segment"),
                s.get("start"),
                s.get("end"),
                s.get("clip_path"),
                s.get("contact_sheet"),
                s.get("download_returncode"),
            ]
        )
    append_sheet(wb, "Video Segments", ["Label", "Video ID", "Segment", "Start", "End", "Clip Path", "Contact Sheet", "Download Return Code"], segment_rows)

    design_rows = [
        ["保留", "持续获得感", "经验/宝箱/金币/升级必须高频出现", "第一局 10 分钟内给出至少 6 次奖励选择"],
        ["保留", "构筑破局感", "玩家要看到从小效果到荒谬组合的曲线", "第 7 分钟出现一次强爆发"],
        ["保留", "低摩擦再开局", "失败也要带回资源和目标", "结算页展示下一局可变强的明确入口"],
        ["改进", "前期重复", "给 2-3 条有效路线，而不是只按数字升序", "初始角色拥有可分叉卡组"],
        ["改进", "UI 可读性", "连击、伤害、关键词和奖励未来收益要可见", "卡牌 hover/预览显示下一张增益"],
        ["突破", "原创世界观绑定机制", "卡牌不是旧武器换名，而是世界观动作", "推荐 Organ Fleet 生理链路"],
        ["风险", "IP 红利缺失", "没有 Vampire Survivors 品牌，玩家耐心更低", "更早展示爆发和差异化"],
    ]
    append_sheet(wb, "Design Implications", ["Type", "Point", "Reason", "Prototype Requirement"], design_rows)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT.resolve())


if __name__ == "__main__":
    main()
